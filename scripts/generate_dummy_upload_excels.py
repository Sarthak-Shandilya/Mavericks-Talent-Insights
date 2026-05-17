"""Generate sample .xlsx files for upload ingestion (human-friendly headers, same as API templates).

Run from repo root:
  python scripts/generate_dummy_upload_excels.py

Output: fixtures/upload_samples/*.xlsx
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from configs.upload_template_headers import (  # noqa: E402
    ASSESSMENTS_HEADERS,
    COMPETENCY_HEADERS,
    STAGES_HEADERS,
    TRAINEE_MASTER_HEADERS,
)

OUT_DIR = ROOT / "fixtures" / "upload_samples"


def _write_sheet(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as e:
        raise PermissionError(
            f"Could not write {path}: file is likely open in Excel or the IDE. Close it and retry."
        ) from e


def _safe_write(path: Path, headers: list[str], rows: list[list[object]]) -> bool:
    try:
        _write_sheet(path, headers, rows)
    except PermissionError as e:
        print(f"SKIP {path.name}: {e}", file=sys.stderr)
        return False
    return True


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # --- TRAINEE_MASTER (same headers as GET /uploads/templates/TRAINEE_MASTER) ---
    trainee_headers = list(TRAINEE_MASTER_HEADERS)
    trainee_rows: list[list[object]] = [
        [
            "EMP-DEMO-001",
            "SUP-10001",
            date(2026, 1, 15),
            "Asha Verma",
            "F",
            "asha.verma.demo@example.com",
            "9876500001",
            "ABC Institute of Tech",
            "Pune",
            "Maharashtra",
            "Mumbai",
            "Mumbai",
            "ACTIVE",
            "JAVA",
            "FOUNDATION",
            "Fresher",
            "Java Backend Track",
            "",
        ],
        [
            "EMP-DEMO-002",
            "SUP-10002",
            date(2026, 1, 20),
            "Rohan Mehta",
            "M",
            "rohan.mehta.demo@example.com",
            "9876500002",
            "XYZ College",
            "Bangalore",
            "Karnataka",
            "Bangalore",
            "Bangalore",
            "ACTIVE",
            "DATA",
            "SPARK",
            "Fresher",
            "Data Engineering Track",
            "",
        ],
    ]
    _safe_write(OUT_DIR / "trainee_master_sample.xlsx", trainee_headers, trainee_rows)

    # --- ASSESSMENTS (long format; employee must exist after trainee upload) ---
    assess_headers = list(ASSESSMENTS_HEADERS)
    assess_rows: list[list[object]] = [
        ["EMP-DEMO-001", "SPARK", "SPARK_P1_A1", 1, 72.5, 100, date(2026, 2, 1), "First attempt"],
        ["EMP-DEMO-001", "SPARK", "SPARK_P1_A2", 1, 81.0, 100, date(2026, 2, 5), "Retry"],
        ["EMP-DEMO-001", "FOUNDATION", "FM1", 1, 88.0, 100, date(2026, 2, 10), None],
        ["EMP-DEMO-002", "SPARK", "SPARK_FINAL", 1, 90.0, 100, date(2026, 2, 3), None],
        ["EMP-DEMO-002", "TECHNICAL", "SQL", 1, 76.0, 100, date(2026, 2, 12), None],
    ]
    _safe_write(OUT_DIR / "assessments_sample.xlsx", assess_headers, assess_rows)

    # --- STAGES ---
    stage_headers = list(STAGES_HEADERS)
    stage_rows: list[list[object]] = [
        ["EMP-DEMO-001", "SPARK", "COMPLETED", 85.0, 2, date(2026, 2, 6)],
        ["EMP-DEMO-001", "FOUNDATION", "PENDING", None, 0, None],
        ["EMP-DEMO-002", "SPARK", "COMPLETED", 90.0, 1, date(2026, 2, 4)],
    ]
    _safe_write(OUT_DIR / "stages_sample.xlsx", stage_headers, stage_rows)

    # --- COMPETENCY ---
    comp_headers = list(COMPETENCY_HEADERS)
    comp_rows: list[list[object]] = [
        ["EMP-DEMO-001", "Java Backend Track", "IN_PROGRESS", "INTERMEDIATE", "No", None],
        ["EMP-DEMO-002", "Data Engineering Track", "NOT_STARTED", "BEGINNER", "No", None],
    ]
    _safe_write(OUT_DIR / "competency_sample.xlsx", comp_headers, comp_rows)

    print("Wrote:")
    for p in sorted(OUT_DIR.glob("*.xlsx")):
        print(" ", p.relative_to(ROOT))


if __name__ == "__main__":
    main()
