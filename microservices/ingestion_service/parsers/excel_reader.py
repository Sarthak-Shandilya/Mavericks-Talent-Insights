from __future__ import annotations

import logging
import re
from collections.abc import Iterator
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _normalize_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed if trimmed else None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return value


def count_data_rows(file_bytes: bytes) -> int:
    """Count non-empty data rows (excludes header) for progress denominator."""
    count = 0
    for _row_number, _row in iter_rows(file_bytes):
        count += 1
    return count


def iter_rows(file_bytes: bytes) -> Iterator[tuple[int, dict[str, object]]]:
    logger.info("excel_reader: load_workbook read_only bytes=%s", len(file_bytes))
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        logger.warning("excel_reader: empty sheet (no header row)")
        return
    headers = [_normalize_header(col) for col in header_row]
    logger.info("excel_reader: headers count=%s sample=%s", len(headers), headers[:8])
    for index, values in enumerate(rows_iter, start=2):
        row = {
            headers[i]: _normalize_value(values[i]) if i < len(values) else None
            for i in range(len(headers))
            if headers[i]
        }
        if all(v is None for v in row.values()):
            continue
        yield index, row
